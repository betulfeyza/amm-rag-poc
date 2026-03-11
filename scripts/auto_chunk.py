"""
Otomatik PDF Parçalama Scripti
-------------------------------
İşlenen sayfalar:
  - 06___086.PDF  → Sayfa 9  (PDF index: 8)
  - 20___086.PDF  → Sayfa 50 (PDF index: 49)
  - 20___086.PDF  → Sayfa 59 (PDF index: 58)

Çıktı: Otomatik_Parcalama.xlsx
"""

import re
import sys
from pathlib import Path

import camelot
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
RAW_DIR = BASE_DIR / "data" / "raw" / "amm"
OUTPUT_FILE = BASE_DIR / "Otomatik_Parcalama.xlsx"

# PDF adı  →  işlenecek sayfa indeksleri (0-tabanlı)
PDF_PAGES: dict[str, list[int]] = {
    "06___086.PDF": [8],       # Sayfa 9
    "20___086.PDF": [49, 58],  # Sayfa 50, 59
}

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
WARNING_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
NOTE_FILL    = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
TASK_FILL    = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
HEADING_FILL = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
HEADER_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def detect_footer_page(page: fitz.Page) -> str | None:
    """Footer'daki 'Page XXX' formatındaki sayfa numarasını bul (sayfanın alt %12'si)."""
    h = page.rect.height
    footer_rect = fitz.Rect(0, h * 0.88, page.rect.width, h)
    footer_text = page.get_text("text", clip=footer_rect).strip()
    # Boeing AMM format: "Page 201", "Page 211" etc.
    m = re.search(r"Page\s+(\d{1,4})", footer_text, re.IGNORECASE)
    if m:
        return m.group(1)
    # Fallback: any standalone number
    numbers = re.findall(r"(?<![\w-])(\d{1,4})(?![\w-])", footer_text)
    return numbers[0] if numbers else None


def collect_font_sizes(raw_blocks: list[dict]) -> list[float]:
    sizes: list[float] = []
    for b in raw_blocks:
        if b.get("type") == 0:
            for line in b.get("lines", []):
                for span in line.get("spans", []):
                    s = span.get("size", 0)
                    if s:
                        sizes.append(s)
    return sizes


def classify_block(
    text: str,
    font_size: float,
    y_coord: float,
    page_height: float,
    all_sizes: list[float],
) -> str:
    """Font boyutu ve konuma göre blok türünü tahmin et."""
    stripped = text.strip()

    # Header / Footer bölgesi
    if y_coord < page_height * 0.05 or y_coord > page_height * 0.90:
        return "Header/Footer"

    # Tablo içeriği (tab veya pipe yoğunluğu)
    if stripped.count("\t") > 3 or stripped.count("|") > 2:
        return "Tablo"

    # Başlık tespiti: sayfadaki max fontun %90'ı ve ortalamadan %30 büyük
    if all_sizes:
        max_f = max(all_sizes)
        avg_f = sum(all_sizes) / len(all_sizes)
        if font_size >= max_f * 0.90 and font_size > avg_f * 1.30:
            return "Başlık"

    return "Gövde Metni"


def build_marker(text: str) -> str:
    """Blok için özel işaret etiketleri üret."""
    upper = text.upper()
    parts: list[str] = []
    if "WARNING" in upper:
        parts.append("⚠ WARNING")
    stripped_upper = upper.lstrip()
    if stripped_upper.startswith("NOTE"):
        parts.append("📝 NOTE")
    if stripped_upper.startswith("TASK") or stripped_upper.startswith("SUBTASK"):
        parts.append("🔧 TASK/SUBTASK")
    return " | ".join(parts)


# ---------------------------------------------------------------------------
# Core extraction
# ---------------------------------------------------------------------------

def extract_page_blocks(pdf_path: Path, page_idx: int) -> tuple[list[dict], str | None]:
    """PyMuPDF ile bir sayfadan tüm metin bloklarını çıkar."""
    doc = fitz.open(str(pdf_path))
    page = doc[page_idx]
    page_height = page.rect.height

    footer_page = detect_footer_page(page)
    raw_dict = page.get_text("dict")  # "dict" mode is reliable across all PDFs
    raw_blocks = raw_dict.get("blocks", [])
    all_sizes = collect_font_sizes(raw_blocks)

    blocks: list[dict] = []
    for block in raw_blocks:
        if block.get("type") != 0:
            continue  # görsel blokları atla

        spans_text: list[str] = []
        font_sizes: list[float] = []

        for line in block.get("lines", []):
            for span in line.get("spans", []):
                t = span.get("text", "").strip()
                if t:
                    spans_text.append(t)
                s = span.get("size", 0)
                if s:
                    font_sizes.append(s)

        full_text = " ".join(spans_text).strip()
        if not full_text:
            continue

        dominant_font = max(font_sizes) if font_sizes else 0.0
        bbox = block.get("bbox", (0, 0, 0, 0))
        y_coord = bbox[1]

        blocks.append(
            {
                "text": full_text,
                "y_coord": y_coord,
                "font_size": round(dominant_font, 2),
                "block_type": classify_block(
                    full_text, dominant_font, y_coord, page_height, all_sizes
                ),
                "marker": build_marker(full_text),
                "footer_page": footer_page,
            }
        )

    doc.close()
    # Yukarıdan aşağıya sırala
    blocks.sort(key=lambda b: b["y_coord"])
    return blocks, footer_page


def blocks_to_dataframe(
    blocks: list[dict],
    pdf_name: str,
    id_start: int = 1,
    fallback_page_label: str = "?",
) -> pd.DataFrame:
    rows = []
    for i, b in enumerate(blocks):
        rows.append(
            {
                "ID": f"AUTO_{id_start + i:03d}",
                "İçerik": b["text"],
                "Kaynak PDF": pdf_name,
                "Sayfa (Footer)": b["footer_page"] or fallback_page_label,
                "Y Koordinatı": b["y_coord"],
                "Font Boyutu": b["font_size"],
                "Tespit Edilen Tür": b["block_type"],
                "Özel İşaret": b["marker"],
            }
        )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Camelot table extraction
# ---------------------------------------------------------------------------

def extract_camelot_tables(pdf_path: Path, page_idx: int) -> list[pd.DataFrame]:
    """Camelot ile tablo çıkart; önce lattice, olmadıysa stream dene."""
    page_str = str(page_idx + 1)
    result_dfs: list[pd.DataFrame] = []

    for flavor in ("lattice", "stream"):
        try:
            tables = camelot.read_pdf(str(pdf_path), pages=page_str, flavor=flavor)
            for t_idx, table in enumerate(tables):
                df = table.df.copy()
                df.insert(
                    0,
                    "Kaynak",
                    f"{pdf_path.name} | Sayfa {page_idx + 1} | {flavor.capitalize()} Tablo {t_idx + 1}",
                )
                result_dfs.append(df)
            if result_dfs:
                break  # lattice başarılı olduysa stream'i deneme
        except Exception as exc:
            print(f"  [Camelot/{flavor}] {pdf_path.name} sayfa {page_idx+1}: {exc}")

    return result_dfs


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------

def style_sheet(ws) -> None:
    """Sütun genişlikleri, başlık rengi ve satır renklendirme."""
    # Başlık satırı
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Sütun genişlikleri (makul varsayılan)
    col_widths = {1: 12, 2: 60, 3: 20, 4: 16, 5: 14, 6: 14, 7: 18, 8: 20}
    for col_idx, width in col_widths.items():
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width

    # Veri satırı renklendirme
    header = [cell.value for cell in ws[1]]
    try:
        special_col_idx = header.index("Özel İşaret")
        type_col_idx = header.index("Tespit Edilen Tür")
    except ValueError:
        return

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        special_val = (row[special_col_idx].value or "").upper()
        type_val = row[type_col_idx].value or ""

        if "WARNING" in special_val:
            fill = WARNING_FILL
        elif "NOTE" in special_val:
            fill = NOTE_FILL
        elif "TASK" in special_val or "SUBTASK" in special_val:
            fill = TASK_FILL
        elif type_val == "Başlık":
            fill = HEADING_FILL
        else:
            continue

        for cell in row:
            cell.fill = fill


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 60)
    print("Otomatik PDF Parçalama Başlıyor...")
    print("=" * 60)

    # -- Sayfa tanımları ve sheet isimleri -----------------------------------
    # (pdf_name, page_idx, sheet_name, fallback_label)
    page_jobs = [
        ("06___086.PDF", 8,  "06___086 Sayfa9",            "9"),
        ("20___086.PDF", 49, "20___086 Sayfa50",            "50"),
        ("20___086.PDF", 58, "20___086 Sayfa59 (Page 211)", "59"),
    ]

    sheet_dfs: dict[str, pd.DataFrame] = {}
    camelot_dfs: list[pd.DataFrame] = []
    auto_id = 1

    for pdf_name, page_idx, sheet_name, fallback in page_jobs:
        pdf_path = RAW_DIR / pdf_name
        print(f"\n[{pdf_name}] Sayfa {page_idx + 1} işleniyor → '{sheet_name}'")

        # Metin blokları
        blocks, footer = extract_page_blocks(pdf_path, page_idx)
        footer_label = f"Footer: {footer}" if footer else f"PDF-Sayfa {page_idx + 1}"
        print(f"  Tespit edilen footer sayfa: {footer or '(bulunamadı)'}")
        print(f"  Metin bloğu sayısı: {len(blocks)}")

        df = blocks_to_dataframe(blocks, pdf_name, id_start=auto_id, fallback_page_label=fallback)
        sheet_dfs[sheet_name] = df
        auto_id += len(blocks)

        # Camelot tabloları
        print(f"  Camelot tablo araması...")
        tbls = extract_camelot_tables(pdf_path, page_idx)
        print(f"  Camelot bulunan tablo: {len(tbls)}")
        camelot_dfs.extend(tbls)

    # -- Excel yazımı -------------------------------------------------------
    print(f"\nExcel dosyası yazılıyor: {OUTPUT_FILE}")

    with pd.ExcelWriter(str(OUTPUT_FILE), engine="openpyxl") as writer:
        for sheet_name, df in sheet_dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Camelot sheet
        if camelot_dfs:
            camelot_combined = pd.concat(camelot_dfs, ignore_index=True)
            camelot_combined.to_excel(writer, sheet_name="Camelot_Tablolar", index=False)
        else:
            pd.DataFrame(
                {"Bilgi": ["Camelot belirtilen sayfalarda tablo bulamadı."]}
            ).to_excel(writer, sheet_name="Camelot_Tablolar", index=False)

    # -- Stil uygulama -------------------------------------------------------
    wb = load_workbook(str(OUTPUT_FILE))
    for sname in list(sheet_dfs.keys()):
        if sname in wb.sheetnames:
            style_sheet(wb[sname])

    # Camelot sheet başlığını da renklendir
    if "Camelot_Tablolar" in wb.sheetnames:
        ws_cam = wb["Camelot_Tablolar"]
        for cell in ws_cam[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

    wb.save(str(OUTPUT_FILE))
    print(f"\n✅ Tamamlandı! Dosya: {OUTPUT_FILE}")

    # Özet
    print("\n── Özet ──────────────────────────────────────────────")
    for sname, df in sheet_dfs.items():
        total = len(df)
        if total == 0 or "Özel İşaret" not in df.columns:
            print(f"  {sname:<35} | 0 blok")
            continue
        w_cnt = df["Özel İşaret"].str.contains("WARNING", na=False).sum()
        n_cnt = df["Özel İşaret"].str.contains("NOTE", na=False).sum()
        t_cnt = df["Özel İşaret"].str.contains("TASK", na=False).sum()
        h_cnt = (df["Tespit Edilen Tür"] == "Başlık").sum()
        print(
            f"  {sname:<35} | {total:3d} blok | "
            f"{w_cnt} WARNING | {n_cnt} NOTE | {t_cnt} TASK | {h_cnt} Başlık"
        )
    print("──────────────────────────────────────────────────────\n")


if __name__ == "__main__":
    main()
